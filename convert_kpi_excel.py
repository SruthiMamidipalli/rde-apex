import openpyxl
import os

file_path = r"KPI (CRM, DRI, Overview).xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

output = []
output.append("# KPI Reference — CRM, DRI, Overview\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    output.append(f"\n## Worksheet: {sheet_name}\n")
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        output.append("*(empty sheet)*\n")
        continue
    
    # Find header row (first non-empty row)
    header = None
    data_start = 0
    for i, row in enumerate(rows):
        if any(cell is not None for cell in row):
            header = row
            data_start = i + 1
            break
    
    if header is None:
        output.append("*(empty sheet)*\n")
        continue
    
    # Clean header
    header = [str(h) if h is not None else "" for h in header]
    col_count = len(header)
    
    # Build markdown table
    output.append("| " + " | ".join(header) + " |")
    output.append("|" + "|".join(["---"] * col_count) + "|")
    
    for row in rows[data_start:]:
        cells = [str(c) if c is not None else "" for c in row[:col_count]]
        # Escape pipe characters in cells
        cells = [c.replace("|", "\\|") for c in cells]
        output.append("| " + " | ".join(cells) + " |")
    
    output.append("")

result = "\n".join(output)

with open("KPI_CRM_DRI_Overview.md", "w", encoding="utf-8") as f:
    f.write(result)

print(f"Done. {len(wb.sheetnames)} worksheets converted.")
print(f"Sheets: {', '.join(wb.sheetnames)}")
